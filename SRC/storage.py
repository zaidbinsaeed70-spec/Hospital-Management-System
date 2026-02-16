"""Persistence helpers for saving/loading application state."""

from __future__ import annotations

import json
import sys
from pathlib import Path

# Try importing openpyxl
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None
    print("!!! WARNING: 'openpyxl' library is NOT installed. Excel files will be ignored. !!!")
    print("Run: pip install openpyxl")

# Relative imports
from .appointment import Appointment
from .hash_table import HashTable
from .patient import Patient
from .priority_queue import PriorityQueue
from .tree import BST

# --- SMART PATH FINDER ---
# We check two locations:
# 1. The project root (E:\DSA\CEP\) -> Best practice
# 2. The src folder (E:\DSA\CEP\src\) -> Fallback
CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent

DATA_FILE = PROJECT_ROOT / "records.json"

# Logic to find the Excel file
POSSIBLE_PATHS = [
    PROJECT_ROOT / "appointment.xlsx",  # Check outer folder
    CURRENT_DIR / "appointment.xlsx"    # Check inner folder
]

EXCEL_FILE = None
for p in POSSIBLE_PATHS:
    if p.exists():
        EXCEL_FILE = p
        break

# Default to outer folder if neither exists (for creating new file)
if EXCEL_FILE is None:
    EXCEL_FILE = PROJECT_ROOT / "appointment.xlsx"

def _row_to_patient(row_values):
    if not row_values or len(row_values) < 5:
        return None

    # Safe extraction
    def safe_str(val): return str(val).strip() if val else ""

    patient_id = safe_str(row_values[0])
    name = safe_str(row_values[1])
    age_raw = row_values[2]
    gender = safe_str(row_values[3]).upper()
    phone = safe_str(row_values[4])
    notes = safe_str(row_values[5]) if len(row_values) > 5 else ""

    if not patient_id or not name or not gender or not phone:
        return None

    try:
        age = int(age_raw)
    except (ValueError, TypeError):
        return None

    if gender not in ("M", "F", "O"):
        if gender.startswith("M"): gender = "M"
        elif gender.startswith("F"): gender = "F"
        else: return None

    try:
        return Patient(patient_id, name, age, gender, phone, notes)
    except Exception:
        return None


def _load_excel_patients(path: Path | str):
    if openpyxl is None:
        return []

    path = Path(path)
    if not path.exists():
        print(f"--> [DEBUG] Excel file NOT found at: {path}")
        return []

    print(f"--> [DEBUG] Found Excel file at: {path}")

    patients = []
    try:
        wb = load_workbook(filename=path, data_only=True)
        # Look for 'Patients' sheet, fallback to active
        if "Patients" in wb.sheetnames:
            ws = wb["Patients"]
        else:
            ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            p = _row_to_patient(row)
            if p:
                patients.append(p)

        wb.close()
    except Exception as e:
        print(f"--> [ERROR] Failed to read Excel: {e}")
        return []

    return patients


def load_state(path: Path | str = DATA_FILE, excel_path: Path | str = EXCEL_FILE):
    registry = HashTable()
    triage = PriorityQueue()
    schedule = BST()
    appt_seq = 0

    # 1. Load JSON
    path = Path(path)
    raw = {}
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as handle:
                raw = json.load(handle)
        except Exception:
            raw = {}

    for pdata in raw.get("patients", []):
        try:
            patient = Patient.from_dict(pdata)
            registry.put(patient.patient_id, patient)
        except Exception:
            continue

    # 2. Import from Excel
    imported_count = 0
    excel_patients = _load_excel_patients(excel_path)
    for p in excel_patients:
        # If not in registry (JSON), add it.
        # This ensures Excel data is loaded into memory.
        if not registry.contains(p.patient_id):
            registry.put(p.patient_id, p)
            imported_count += 1
        else:
            # If already exists in JSON, we ignore Excel to preserve visit history
            pass

    # 3. Restore Triage
    for entry in raw.get("triage", []):
        try:
            triage.enqueue(entry["priority"], entry.get("payload"))
        except Exception:
            continue

    # 4. Restore Appointments
    for appt_data in raw.get("appointments", []):
        try:
            dt_value = appt_data.get("datetime") or appt_data.get("dt")
            if not dt_value:
                continue
            appt = Appointment(
                appt_data["patient_id"],
                dt_value,
                code=appt_data.get("code"),
            )
            schedule.insert(appt_data["key"], appt)
        except Exception:
            continue

    saved_seq = raw.get("appt_seq", 0)
    if isinstance(saved_seq, int) and saved_seq >= 0:
        appt_seq = saved_seq

    return registry, triage, schedule, appt_seq, imported_count


def save_state(registry: HashTable, triage: PriorityQueue, schedule: BST, appt_seq: int,
               path: Path | str = DATA_FILE, excel_path: Path | str = EXCEL_FILE):
    """Persist data to JSON and generate the multi-sheet Excel file."""
    path = Path(path)

    # Save JSON
    data = {
        "patients": [],
        "triage": [],
        "appointments": [],
        "appt_seq": appt_seq,
    }

    all_patients = []
    for key, patient in registry.items():
        all_patients.append(patient)
        try:
            data["patients"].append(patient.to_dict())
        except Exception:
            continue

    data["patients"].sort(key=lambda item: item["patient_id"])
    all_patients.sort(key=lambda p: p.patient_id)

    for entry in triage.to_list():
        data["triage"].append(entry)

    for key, appt in schedule.items():
        try:
            data["appointments"].append(
                {
                    "key": key,
                    "code": appt.code,
                    "patient_id": appt.patient_id,
                    "datetime": appt.datetime_str(),
                }
            )
        except Exception:
            continue

    try:
        with path.open("w", encoding="utf-8") as handle:
            json.dump(data, handle, indent=2)
    except Exception:
        pass

    # Save Excel
    _save_full_excel(all_patients, schedule, triage, registry, excel_path)


def _save_full_excel(patients_list, schedule: BST, triage: PriorityQueue, registry: HashTable, path: Path | str):
    if openpyxl is None:
        return

    path = Path(path)
    try:
        wb = Workbook()

        # SHEET 1: Patients
        ws1 = wb.active
        ws1.title = "Patients"
        ws1.append(["Patient ID", "Name", "Age", "Gender", "Phone", "Medical Notes"])

        for p in patients_list:
            ws1.append([
                p.patient_id, p.name, p.age, p.gender, p.phone, p.medical_notes
            ])

        # SHEET 2: Appointments
        ws2 = wb.create_sheet(title="Appointments")
        ws2.append(["Appt Code", "Date & Time", "Patient ID", "Patient Name", "Status"])

        # Iterate through BST items
        def _collect(node):
            if node is None: return
            _collect(node.left)
            appt = node.value
            p_name = "Unknown"
            p = registry.get(appt.patient_id)
            if p: p_name = p.name
            ws2.append([appt.code, appt.datetime_str(), appt.patient_id, p_name, "Scheduled"])
            _collect(node.right)

        _collect(schedule.root)

        # SHEET 3: Emergencies
        ws3 = wb.create_sheet(title="Emergencies")
        ws3.append(["Priority", "Details"])

        for item in triage.to_list():
            ws3.append([item['priority'], item['payload']])

        wb.save(path)
    except Exception as e:
        print(f"Error saving to Excel: {e}")